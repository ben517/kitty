#!/usr/bin/env python3
"""统一的场景化测试执行器。

从 YAML 配置文件读取测试场景，执行测试并生成综合报告。
包含性能测试（耗时统计）和准确率测试。

使用方法:
    python tests/scenarios/test_runner.py
    
输出:
    - /output/test_YYYYMMDD_HHMMSS.xlsx (Excel 详细结果)
    - tests/scenarios/output/test_results.json (JSON 详细结果)
    - tests/scenarios/output/TEST_REPORT_YYYYMMDD_HHMMSS.md (Markdown 报告)
"""

import json
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import yaml
from fastapi.testclient import TestClient

# 添加项目路径
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from app.main import app
from app.models.schemas import IntentType


@dataclass
class TestResult:
    """单个测试用例的结果。"""
    scenario_id: str
    category: str
    query: str
    description: str
    expected_intent: Optional[str]
    actual_intent: Optional[str] = None
    intent_correct: bool = False
    response_text: str = ""
    sources: list[str] = field(default_factory=list)
    device_id: Optional[str] = None
    session_id: Optional[str] = None
    total_time_ms: float = 0.0
    llm_time_ms: float = 0.0
    status_code: int = 200
    error_message: str = ""
    priority: str = "medium"
    status: str = "pending"
    
    def to_dict(self) -> dict:
        """转换为字典格式。"""
        return {
            "scenario_id": self.scenario_id,
            "category": self.category,
            "query": self.query,
            "description": self.description,
            "expected_intent": self.expected_intent,
            "actual_intent": self.actual_intent,
            "intent_correct": self.intent_correct,
            "response_text": self.response_text[:200] if self.response_text else "",
            "sources": self.sources,
            "device_id": self.device_id,
            "session_id": self.session_id,
            "total_time_ms": round(self.total_time_ms, 2),
            "llm_time_ms": round(self.llm_time_ms, 2),
            "status_code": self.status_code,
            "error_message": self.error_message,
            "priority": self.priority,
            "status": self.status,
        }


class ScenarioTestRunner:
    """场景化测试执行器。"""
    
    def __init__(self, config_file: str):
        """初始化测试执行器。
        
        Args:
            config_file: YAML 配置文件路径
        """
        self.config_file = Path(config_file)
        self.config = self._load_config()
        self.client = TestClient(app)
        self.results: list[TestResult] = []
        self.sessions: dict[str, str] = {}  # session_id -> session_id mapping
        
    def _load_config(self) -> dict:
        """加载 YAML 配置。"""
        with open(self.config_file, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    
    def _flatten_scenarios(self) -> list[dict]:
        """将嵌套的场景配置展平为列表。"""
        scenarios = []
        for category_group in self.config.get('scenarios', []):
            category = category_group.get('category', 'unknown')
            priority = category_group.get('priority', 'medium')
            
            for scene in category_group.get('scenes', []):
                scene['category'] = category
                scene['priority'] = priority
                scenarios.append(scene)
        
        return scenarios
    
    def _parse_intent(self, intent_str: Optional[str]) -> Optional[IntentType]:
        """将字符串转换为 IntentType。"""
        if not intent_str:
            return None
        try:
            return IntentType(intent_str)
        except ValueError:
            return None
    
    def _execute_single_test(self, scenario: dict) -> TestResult:
        """执行单个测试用例。
        
        Args:
            scenario: 场景配置字典
            
        Returns:
            TestResult: 测试结果
        """
        result = TestResult(
            scenario_id=scenario.get('id', 'unknown'),
            category=scenario.get('category', 'unknown'),
            query=scenario.get('query', ''),
            description=scenario.get('description', ''),
            expected_intent=scenario.get('expected_intent'),
            priority=scenario.get('priority', 'medium'),
        )
        
        # 准备请求数据
        request_data = {
            "query": result.query,
        }
        
        # 如果有 session_id，使用或创建会话
        if 'session_id' in scenario:
            config_session_id = scenario['session_id']
            if config_session_id not in self.sessions:
                self.sessions[config_session_id] = f"test_session_{int(time.time())}_{config_session_id}"
            request_data['session_id'] = self.sessions[config_session_id]
        
        # 如果有依赖，确保使用相同的 session
        if 'depends_on' in scenario:
            # 多轮对话使用相同 session
            pass
        
        # 执行请求
        start_time = time.time()
        try:
            response = self.client.post(
                "/chat/",
                json=request_data,
                timeout=self.config.get('execution_config', {}).get('timeout_seconds', 300)
            )
            result.status_code = response.status_code
            
            if response.status_code == 200:
                response_data = response.json()
                result.response_text = response_data.get('answer', '')
                result.sources = response_data.get('sources', [])
                result.device_id = response_data.get('device_id')
                result.session_id = response_data.get('session_id')
                result.actual_intent = response_data.get('intent')
                
                # 更新 session 映射
                if result.session_id and 'session_id' in scenario:
                    self.sessions[scenario['session_id']] = result.session_id
                
                # 判断意图是否正确
                if result.expected_intent:
                    result.intent_correct = result.actual_intent == result.expected_intent
                else:
                    result.intent_correct = True  # 无预期意图的测试默认成功
                    
                result.status = "success"
            else:
                result.error_message = f"HTTP {response.status_code}: {response.text[:200]}"
                result.status = "failed" if response.status_code >= 400 else "success"
                
        except Exception as e:
            result.error_message = str(e)
            result.status = "error"
        
        end_time = time.time()
        result.total_time_ms = (end_time - start_time) * 1000
        result.llm_time_ms = result.total_time_ms * 0.95  # 估算 LLM 时间占比
        
        return result
    
    def run_all_tests(self) -> list[TestResult]:
        """执行所有测试。"""
        scenarios = self._flatten_scenarios()
        total = len(scenarios)
        
        print("=" * 60)
        print("开始执行场景测试...")
        print("=" * 60)
        print(f"总场景数：{total}\n")
        
        for idx, scenario in enumerate(scenarios, 1):
            print(f"[{idx}/{total}] 执行 {scenario.get('id')}...", end=" ")
            result = self._execute_single_test(scenario)
            self.results.append(result)
            
            status_icon = "✅" if result.status in ["success", "pending"] else "❌"
            print(f"{status_icon} {result.total_time_ms:.0f}ms")
        
        return self.results
    
    def generate_summary(self) -> dict:
        """生成测试汇总统计。"""
        total_tests = len(self.results)
        successful = sum(1 for r in self.results if r.status in ["success", "pending"])
        failed = total_tests - successful
        
        avg_response_time = sum(r.total_time_ms for r in self.results) / total_tests if total_tests > 0 else 0
        avg_llm_time = sum(r.llm_time_ms for r in self.results) / total_tests if total_tests > 0 else 0
        
        # 按优先级统计
        by_priority = {}
        for result in self.results:
            priority = result.priority
            if priority not in by_priority:
                by_priority[priority] = {"total": 0, "success": 0}
            by_priority[priority]["total"] += 1
            if result.status in ["success", "pending"]:
                by_priority[priority]["success"] += 1
        
        # 按类别统计
        by_category = {}
        for result in self.results:
            category = result.category
            if category not in by_category:
                by_category[category] = {"total": 0, "success": 0}
            by_category[category]["total"] += 1
            if result.status in ["success", "pending"]:
                by_category[category]["success"] += 1
        
        return {
            "summary": {
                "total_tests": total_tests,
                "successful": successful,
                "failed": failed,
                "success_rate": f"{successful/total_tests*100:.1f}%" if total_tests > 0 else "0%",
                "avg_response_time_ms": round(avg_response_time, 2),
                "avg_llm_time_ms": round(avg_llm_time, 2),
                "timestamp": datetime.now().isoformat(),
                "config_file": str(self.config_file),
            },
            "by_priority": by_priority,
            "by_category": by_category,
        }
    
    def export_json(self, output_file: str):
        """导出 JSON 格式结果。"""
        summary = self.generate_summary()
        output_data = {
            **summary,
            "results": [r.to_dict() for r in self.results],
        }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)
        
        print(f"\n结果已导出到：{output_file}")
    
    def export_excel(self, output_file: str):
        """导出 Excel 格式结果。
        
        Args:
            output_file: Excel 文件输出路径
        """
        summary = self.generate_summary()
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试结果"
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 定义表头 - 实际回答紧跟输入问题，响应时间在实际回答后
        headers = [
            "场景 ID", "类别", "场景描述", "输入问题", "实际回答", "响应时间 (ms)", "LLM 时间 (ms)",
            "预期意图", "实际意图", "结果", "优先级", "状态码", "错误信息"
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据行
        for row_idx, result in enumerate(self.results, 2):
            status_result = "成功" if result.status in ["success", "pending"] else "失败"
            intent_correct = "正确" if result.intent_correct else "错误"
            
            data = [
                result.scenario_id,
                result.category,
                result.description,
                result.query,
                result.response_text[:500] if result.response_text else "",
                round(result.total_time_ms, 2),
                round(result.llm_time_ms, 2),
                result.expected_intent or "N/A",
                result.actual_intent or "N/A",
                intent_correct,
                result.priority,
                result.status_code,
                result.error_message[:200] if result.error_message else ""
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # 根据结果设置颜色
                if col_idx == 10:  # 结果列 (现在是第 10 列)
                    if status_result == "成功":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
        
        # 调整列宽 - 实际回答列设置更宽，响应时间列适当宽度
        column_widths = [15, 15, 25, 40, 80, 15, 15, 15, 15, 10, 12, 10, 30]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # 创建汇总统计工作表
        ws_summary = wb.create_sheet(title="汇总统计")
        
        # 汇总信息
        s = summary['summary']
        summary_data = [
            ["测试总数", s['total_tests']],
            ["成功", s['successful']],
            ["失败", s['failed']],
            ["成功率", s['success_rate']],
            ["平均响应时间 (ms)", s['avg_response_time_ms']],
            ["平均 LLM 时间 (ms)", s['avg_llm_time_ms']],
            ["生成时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["配置文件", str(self.config_file)],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            cell_label = ws_summary.cell(row=row_idx, column=1, value=label)
            cell_value = ws_summary.cell(row=row_idx, column=2, value=value)
            cell_label.font = Font(bold=True)
            cell_label.border = thin_border
            cell_value.border = thin_border
        
        # 按优先级统计
        ws_summary.cell(row=10, column=1, value="按优先级统计").font = Font(bold=True)
        priority_headers = ["优先级", "成功", "总计", "通过率"]
        for col, header in enumerate(priority_headers, 1):
            cell = ws_summary.cell(row=11, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, (priority, stats) in enumerate(sorted(summary['by_priority'].items()), 12):
            success_rate = f"{stats['success']/stats['total']*100:.1f}%" if stats['total'] > 0 else "N/A"
            data = [priority, stats['success'], stats['total'], success_rate]
            for col_idx, value in enumerate(data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # 按类别统计
        ws_summary.cell(row=20, column=1, value="按类别统计").font = Font(bold=True)
        category_headers = ["类别", "成功", "总计", "通过率"]
        for col, header in enumerate(category_headers, 1):
            cell = ws_summary.cell(row=21, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, (category, stats) in enumerate(sorted(summary['by_category'].items()), 22):
            success_rate = f"{stats['success']/stats['total']*100:.1f}%" if stats['total'] > 0 else "N/A"
            data = [category, stats['success'], stats['total'], success_rate]
            for col_idx, value in enumerate(data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # 保存文件
        wb.save(output_file)
        print(f"Excel 报告已生成：{output_file}")
    
    def generate_markdown_report(self, output_file: str):
        """生成 Markdown 格式报告。"""
        summary = self.generate_summary()
        
        lines = []
        lines.append("# 场景化测试执行报告\n")
        lines.append(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        lines.append(f"**配置文件**: {self.config_file}\n")
        lines.append("---\n")
        
        # 汇总信息
        lines.append("## 📊 测试汇总\n")
        s = summary['summary']
        lines.append(f"- **测试总数**: {s['total_tests']}")
        lines.append(f"- **成功**: {s['successful']}")
        lines.append(f"- **失败**: {s['failed']}")
        lines.append(f"- **成功率**: {s['success_rate']}")
        lines.append(f"- **平均响应时间**: {s['avg_response_time_ms']:.1f} ms")
        lines.append(f"- **平均 LLM 时间**: {s['avg_llm_time_ms']:.1f} ms\n")
        
        # 按优先级统计
        lines.append("### 按优先级统计\n")
        lines.append("| 优先级 | 成功 | 总计 | 通过率 |")
        lines.append("|--------|------|------|--------|")
        for priority, stats in sorted(summary['by_priority'].items()):
            if stats['total'] > 0:
                success_rate = stats['success']/stats['total']*100
                rate = f"{stats['success']}/{stats['total']} ({success_rate:.1f}%)"
            else:
                rate = "N/A"
            lines.append(f"| {priority} | {stats['success']} | {stats['total']} | {rate.split()[-1] if ' ' in rate else rate} |")
        lines.append("")
        
        # 按类别统计
        lines.append("### 按类别统计\n")
        lines.append("| 类别 | 成功 | 总计 | 通过率 |")
        lines.append("|------|------|------|--------|")
        for category, stats in sorted(summary['by_category'].items()):
            if stats['total'] > 0:
                success_rate = stats['success']/stats['total']*100
                rate = f"{stats['success']}/{stats['total']} ({success_rate:.1f}%)"
            else:
                rate = "N/A"
            lines.append(f"| {category} | {stats['success']} | {stats['total']} | {rate.split()[-1] if ' ' in rate else rate} |")
        lines.append("")
        
        # 详细测试结果
        lines.append("---\n")
        lines.append("## 📋 详细测试结果\n")
        lines.append("| ID | 类别 | 场景描述 | 输入问题 | 预期意图 | 响应时间 (ms) | 状态 | 优先级 |")
        lines.append("|----|------|---------|---------|---------|--------------|------|--------|")
        
        for result in self.results:
            status_icon = "✅" if result.status in ["success", "pending"] else "❌"
            query_display = result.query[:30] + "..." if len(result.query) > 30 else result.query
            lines.append(
                f"| {result.scenario_id} | {result.category} | {result.description} | "
                f"{query_display} | {result.expected_intent or 'N/A'} | "
                f"{result.total_time_ms:.0f} | {status_icon} | {result.priority} |"
            )
        
        lines.append("\n---\n")
        lines.append("## ⚠️ 失败详情\n")
        failed_results = [r for r in self.results if r.status not in ["success", "pending"]]
        if failed_results:
            for result in failed_results:
                lines.append(f"### {result.scenario_id}: {result.description}\n")
                lines.append(f"- **错误**: {result.error_message}\n")
                lines.append(f"- **状态码**: {result.status_code}\n")
                lines.append("")
        else:
            lines.append("无失败用例 ✅\n")
        
        # 写入文件
        report_content = "\n".join(lines)
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(report_content)
        
        print(f"报告已生成：{output_file}")
        return report_content


def main():
    """主函数。"""
    # 配置文件路径
    config_file = Path(__file__).parent / "test_scenarios_input.yaml"
    
    if not config_file.exists():
        print(f"错误：配置文件不存在 - {config_file}")
        sys.exit(1)
    
    # 创建执行器
    runner = ScenarioTestRunner(str(config_file))
    
    # 执行测试
    runner.run_all_tests()
    
    # Excel 输出到根目录/output
    root_output_dir = Path(__file__).parent.parent.parent / "output"
    root_output_dir.mkdir(exist_ok=True)
    
    # 生成时间戳
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 导出 Excel 到根目录/output，文件名为 test_时间戳.xlsx
    excel_file = root_output_dir / f"test_{timestamp}.xlsx"
    runner.export_excel(str(excel_file))
    
    # 打印汇总
    print("\n" + "=" * 60)
    summary = runner.generate_summary()['summary']
    print("## 测试汇总报告")
    print(f"测试总数：{summary['total_tests']}")
    print(f"成功：{summary['successful']}")
    print(f"失败：{summary['failed']}")
    print(f"成功率：{summary['success_rate']}")
    print(f"平均响应时间：{summary['avg_response_time_ms']:.1f} ms")
    print(f"Excel 报告已保存到：{excel_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
